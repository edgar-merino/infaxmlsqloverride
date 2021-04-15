#!/usr/bin/env python
# -*- coding: utf-8 -*-

'''For a given IPC exported file, extract those nodes when a sql overrides 
have been specified an write it to an excel file with the name of the last 
forder read from the input file. 
'''

# standard libraries
from string import Template
import sys, os
# third party libraries
import lxml.etree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Long column width
LONG_COLUMN=90

def print_single_node(node, attr=None):
	'''Gets the NAME attr for a given node'''
	if node is None:
		return ''
	if attr is None:
		attr='$NAME'
	s=Template(attr)
	return s.safe_substitute(node.attrib)


def expand_node(node, format=None):
	'''For a given node obtains its hierachy'''
	result=[]
	curnode=node
	format=format if format is not None else []
	while curnode is not None:
		tag=curnode.tag
		attr=None
		for item in format:
			attr=item[tag] if tag in item else None
		result.insert(0, print_single_node(curnode,attr) )
		parent=curnode.getparent()
		if parent is not None and parent.tag == 'REPOSITORY':
			break
		curnode=parent
	return result

def as_text(value):
	'''retur the specified value as a string'''
	if value is None:
		return ""
	return str(value)

def export_excel(data, filename, sheetname=None):
	'''Given a dataset (as a list of lists), save it to the specified file
	with the specified sheetname as an Excel file'''
	if len(data)==0:
		return
	workbook = Workbook()
	sheet = workbook.active
	if sheetname is not None:
		sheet.title=sheetname
	for row in data:
		sheet.append(row)
	# freeze first column
	sheet.freeze_panes="A2"
	# add filter
	sheet.auto_filter.ref=sheet.dimensions
	# autosize (or kind of) columns
	for column_cells in sheet.columns:
		length = max(len(as_text(cell.value)) for cell in column_cells)+5
		length = LONG_COLUMN if length > LONG_COLUMN else length
		sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
	# set column height
	for x in range(sheet.min_row+1,sheet.max_row+1):
		sheet.row_dimensions[x].height = 60
	# wrap text for long columns
	for column_cells in sheet.columns:
		# look for columns larger than LONG_COLUMN to wrap it out
		if sheet.column_dimensions[get_column_letter(column_cells[0].column)].width >= LONG_COLUMN:
			col=column_cells[0].column
			# set wrap text for each colum
			for rows in sheet.iter_rows(min_row=sheet.min_row+1, max_row=sheet.max_row, min_col=col, max_col=col):
				for cell in rows:
					cell.alignment = Alignment(wrapText=True)
	# save it
	workbook.save(filename=filename)

def extract_sql_overrides(infile):
	'''Parse the given file to extract those nodes with a sql query override specified'''
	if not os.path.isfile(infile):
		raise Exception(f"File {infile} does not exists, please verify")

	print(f"Parsing {infile} ...")

	root = ET.parse(infile)
	override=root.xpath('//TABLEATTRIBUTE[@NAME="Sql Query"]')
	result=[['Folder','Mapping','Source Qualifier','Transformation type','SQL statement']]
	filename=''

	print(f"Processing override SQL statements ...")
	
	for node in override:
		sqlstatement=node.attrib['VALUE']
		if sqlstatement != '':
			# NOTE: SQL overrided not required
			exp=expand_node(node)
			exp.append(sqlstatement)
			result.append(exp)
			# TODO: Set filename as the same as infile (without extension)
			filename=exp[0]

	if filename != '':
		print(f"Generating excel file: {filename}.xlsx ...")
		export_excel(result, filename+'.xlsx')
	print(f"Done")

if __name__ == '__main__':
	# TODO: Process multiple files
	if len(sys.argv) < 2:
		print("No file provided, please verify.")
		sys.exit(1)

	infile=sys.argv[1]

	extract_sql_overrides(infile)
